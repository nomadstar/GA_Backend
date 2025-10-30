#!/usr/bin/env python3
"""
Herramienta de verificación del mapeo maestro.
Ejecutar: python3 verify_mapeo.py

Valida que los 3 archivos Excel se pueden leer correctamente
y genera reporte de cobertura.
"""

import openpyxl
import pandas as pd
from pathlib import Path
from collections import defaultdict
import unicodedata
import re


def normalize_name(s: str) -> str:
    """Normalizar nombre igual a la función Rust"""
    # Remover acentos
    s = ''.join(
        c for c in unicodedata.normalize('NFD', s)
        if unicodedata.category(c) != 'Mn'
    )
    # Minúsculas, mantener solo alfanuméricos y espacios
    s = re.sub(r'[^a-z0-9\s]', ' ', s.lower())
    # Colapsar espacios
    s = ' '.join(s.split())
    return s


def leer_malla2020(path: str) -> dict:
    """Leer Malla2020.xlsx"""
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb['Malla2020']
        
        resultados = {}
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            nombre = str(row[0] or "").strip()
            id_str = str(row[1] or "").strip()
            
            if nombre and id_str.isdigit():
                nombre_norm = normalize_name(nombre)
                resultados[nombre_norm] = {
                    'nombre': nombre,
                    'id': int(id_str),
                }
        
        return resultados
    except Exception as e:
        print(f"❌ Error leyendo Malla2020: {e}")
        return {}


def leer_oa2024(path: str) -> dict:
    """Leer OA2024.xlsx"""
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        
        resultados = {}
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            codigo = str(row[1] or "").strip()
            nombre = str(row[2] or "").strip()
            
            if codigo and nombre:
                nombre_norm = normalize_name(nombre)
                if nombre_norm not in resultados:
                    resultados[nombre_norm] = {
                        'nombre': nombre,
                        'codigo': codigo,
                    }
        
        return resultados
    except Exception as e:
        print(f"❌ Error leyendo OA2024: {e}")
        return {}


def leer_pa2025(path: str) -> dict:
    """Leer PA2025-1.xlsx"""
    try:
        df = pd.read_excel(path)
        resultados = {}
        
        for _, row in df.iterrows():
            codigo = str(row.get('Código Asignatura', '')).strip()
            nombre = str(row.get('Nombre', '')).strip()
            porcentaje = row.get('Porcentaje Aprobado')
            es_electivo = bool(row.get('Electivo', False))
            
            if codigo and nombre:
                nombre_norm = normalize_name(nombre)
                if nombre_norm not in resultados:
                    resultados[nombre_norm] = {
                        'nombre': nombre,
                        'codigo': codigo,
                        'porcentaje': porcentaje,
                        'es_electivo': es_electivo,
                    }
        
        return resultados
    except Exception as e:
        print(f"❌ Error leyendo PA2025-1: {e}")
        return {}


def main():
    data_dir = Path("src/datafiles")
    
    print("=" * 80)
    print("📊 VERIFICADOR DE MAPEO MAESTRO")
    print("=" * 80)
    
    # Leer archivos
    print("\n📖 Leyendo archivos Excel...")
    malla = leer_malla2020(str(data_dir / "malla2020.xlsx"))
    oa2024 = leer_oa2024(str(data_dir / "OA2024.xlsx"))
    pa2025 = leer_pa2025(str(data_dir / "PA2025-1.xlsx"))
    
    print(f"  ✓ Malla2020: {len(malla)} asignaturas")
    print(f"  ✓ OA2024: {len(oa2024)} asignaturas únicas")
    print(f"  ✓ PA2025-1: {len(pa2025)} asignaturas únicas")
    
    # Análisis de cobertura
    print("\n" + "=" * 80)
    print("📈 ANÁLISIS DE COBERTURA")
    print("=" * 80)
    
    # Malla en OA2024
    malla_en_oa = sum(1 for m in malla if m in oa2024)
    malla_no_en_oa = len(malla) - malla_en_oa
    print(f"\nMalla2020 → OA2024:")
    print(f"  ✓ {malla_en_oa}/{len(malla)} encontrados en OA2024")
    print(f"  ✗ {malla_no_en_oa} NO encontrados")
    
    # Malla en PA2025-1
    malla_en_pa = sum(1 for m in malla if m in pa2025)
    malla_no_en_pa = len(malla) - malla_en_pa
    print(f"\nMalla2020 → PA2025-1:")
    print(f"  ✓ {malla_en_pa}/{len(malla)} encontrados en PA2025-1")
    print(f"  ✗ {malla_no_en_pa} NO encontrados")
    
    # OA2024 en PA2025-1 (importante!)
    oa_en_pa = 0
    oa_no_en_pa = []
    for norm_name, oa_data in oa2024.items():
        if norm_name in pa2025:
            oa_en_pa += 1
        else:
            oa_no_en_pa.append((norm_name, oa_data['nombre'], oa_data['codigo']))
    
    print(f"\nOA2024 → PA2025-1 (CRÍTICO para schedule solver):")
    print(f"  ✓ {oa_en_pa}/{len(oa2024)} códigos de OA2024 tienen ofertas en PA2025-1")
    print(f"  ✗ {len(oa_no_en_pa)} NO tienen secciones en enero 2025:")
    for norm, nombre, cod in sorted(oa_no_en_pa)[:5]:
        print(f"    - {cod} ({nombre})")
    if len(oa_no_en_pa) > 5:
        print(f"    ... y {len(oa_no_en_pa)-5} más")
    
    # Búsqueda de cambios de código (el problema)
    print("\n" + "=" * 80)
    print("🔍 DETECCIÓN DE CAMBIOS DE CÓDIGO (Problema descubierto)")
    print("=" * 80)
    
    cambios_codigo = []
    for norm_name in malla:
        if norm_name in oa2024 and norm_name in pa2025:
            cod_oa = oa2024[norm_name].get('codigo')
            cod_pa = pa2025[norm_name].get('codigo')
            if cod_oa != cod_pa:
                cambios_codigo.append((
                    norm_name,
                    malla[norm_name]['nombre'],
                    cod_oa,
                    cod_pa
                ))
    
    if cambios_codigo:
        print(f"\n⚠️  {len(cambios_codigo)} asignaturas tienen CÓDIGOS DIFERENTES entre años:")
        for norm, nombre, cod_oa, cod_pa in sorted(cambios_codigo)[:10]:
            print(f"  • {nombre}")
            print(f"    OA2024:  {cod_oa}")
            print(f"    PA2025:  {cod_pa}")
            print()
        
        if len(cambios_codigo) > 10:
            print(f"  ... y {len(cambios_codigo)-10} más")
    else:
        print("✅ Todos los códigos coinciden (raro, esperabas cambios)")
    
    # Resumen final
    print("\n" + "=" * 80)
    print("✅ MAPEO MAESTRO VIABILIDAD")
    print("=" * 80)
    
    cobertura_total = (malla_en_oa / len(malla)) * 100 if malla else 0
    cobertura_pa = (malla_en_pa / len(malla)) * 100 if malla else 0
    
    print(f"\n📊 Estadísticas:")
    print(f"  • Cobertura Malla → OA2024: {cobertura_total:.1f}%")
    print(f"  • Cobertura Malla → PA2025-1: {cobertura_pa:.1f}%")
    print(f"  • Asignaturas con cambio de código: {len(cambios_codigo)}")
    
    if cobertura_pa >= 90:
        print(f"\n✅ La estrategia de NOMBRE como clave universal es VIABLE")
    else:
        print(f"\n⚠️  Cobertura baja, revisar nombres normalizados")
    
    print()


if __name__ == "__main__":
    main()
